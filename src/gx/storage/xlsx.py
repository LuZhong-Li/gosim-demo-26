"""基于 openpyxl 的本地 xlsx 存储实现。

首行作为表头；写操作自动加内存锁并在锁内立即 save()。
参见 docs/plans/02-核心模块设计.md 3.1。
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from constants import (
    AUDIT_LOG,
    ERR_AUDIT_WRITE,
    ERR_STORAGE_FILE_NOT_FOUND,
    ERR_STORAGE_IO,
    ERR_STORAGE_ROW,
    ERR_STORAGE_SHEET,
)
from errors import GXError
from gx.storage.base import BaseStorage
from gx.storage.lock import MemoryLock


class LocalXlsxStorage(BaseStorage):
    """本地 xlsx 存储：唯一接触 openpyxl 的地方。"""

    def __init__(self, path: str) -> None:
        """加载本地 xlsx 工作簿；文件不存在抛 S001，读取失败抛 S005。"""
        self._path = path
        self._lock = MemoryLock()
        if not Path(path).is_file():
            raise GXError(
                ERR_STORAGE_FILE_NOT_FOUND,
                f"工作簿文件不存在: {path}",
                module="storage",
                context={"path": path},
            )
        try:
            self._workbook = load_workbook(path)
        except GXError:
            raise
        except Exception as exc:
            # 损坏文件 / 权限不足等 IO 异常统一转友好错误码 S005
            raise GXError(
                ERR_STORAGE_IO,
                f"工作簿读取失败: {exc}",
                module="storage",
                context={"path": path},
            ) from exc

    @classmethod
    def create_workbook(cls, path: str) -> "LocalXlsxStorage":
        """新建空工作簿并保存，返回对应存储实例（种子脚本用）。"""
        workbook = Workbook()
        try:
            workbook.save(path)
        except OSError as exc:
            raise GXError(
                ERR_STORAGE_IO,
                f"工作簿创建失败: {exc}",
                module="storage",
                context={"path": path},
            ) from exc
        return cls(path)

    def get_sheet(self, sheet_name: str) -> list[dict[str, Any]]:
        """读整表，返回 [{列名: 值}, ...]；表头为第 0 行，数据行从 0 编号。"""
        self._ensure_sheet(sheet_name)
        worksheet = self._workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(header) for header in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]

    def append_row(self, sheet_name: str, row: dict[str, Any]) -> None:
        """向工作表追加一行，键为表头列名；写操作带内存锁并立即落盘。"""
        with self.lock():
            self._ensure_sheet(sheet_name)
            worksheet = self._workbook[sheet_name]
            headers = [str(cell.value) for cell in worksheet[1]]
            worksheet.append([row.get(header) for header in headers])
            self.save()

    def update_row(self, sheet_name: str, row_id: int, data: dict[str, Any]) -> None:
        """更新第 row_id 条数据行（从 0 开始），只更新 data 中出现的列。"""
        with self.lock():
            self._ensure_sheet(sheet_name)
            if sheet_name == AUDIT_LOG:
                raise GXError(
                    ERR_AUDIT_WRITE,
                    "audit_log 表只允许追加，禁止更新/删除",
                    module="storage",
                    context={"sheet_name": sheet_name},
                )
            worksheet = self._workbook[sheet_name]
            data_row_count = worksheet.max_row - 1
            if row_id < 0 or row_id >= data_row_count:
                raise GXError(
                    ERR_STORAGE_ROW,
                    f"数据行不存在或越界: row_id={row_id}",
                    module="storage",
                    context={"sheet_name": sheet_name, "row_id": row_id},
                )
            excel_row = row_id + 2  # 表头为第 0 行，数据行从 0 编号
            headers = [str(cell.value) for cell in worksheet[1]]
            for column_index, header in enumerate(headers, start=1):
                if header in data:
                    worksheet.cell(row=excel_row, column=column_index, value=data[header])
            self.save()

    def add_sheet(self, sheet_name: str, columns: list[str]) -> None:
        """新建工作表，columns 作为首行表头；同名工作表已存在抛 S002。"""
        with self.lock():
            if sheet_name in self._workbook.sheetnames:
                raise GXError(
                    ERR_STORAGE_SHEET,
                    f"工作表已存在: {sheet_name}",
                    module="storage",
                    context={"sheet_name": sheet_name},
                )
            worksheet = self._workbook.create_sheet(title=sheet_name)
            worksheet.append(columns)
            self.save()

    def remove_sheet(self, sheet_name: str) -> None:
        """删除工作表（种子脚本清理 openpyxl 默认工作表用）。"""
        with self.lock():
            self._ensure_sheet(sheet_name)
            if sheet_name == AUDIT_LOG:
                raise GXError(
                    ERR_AUDIT_WRITE,
                    "audit_log 表只允许追加，禁止删除",
                    module="storage",
                    context={"sheet_name": sheet_name},
                )
            self._workbook.remove(self._workbook[sheet_name])
            if self._workbook.sheetnames:
                self.save()

    def save(self) -> None:
        """落盘（写操作后立即调用）；IO 失败抛 S005。"""
        try:
            self._workbook.save(self._path)
        except OSError as exc:
            raise GXError(
                ERR_STORAGE_IO,
                f"工作簿写入失败: {exc}",
                module="storage",
                context={"path": self._path},
            ) from exc

    def lock(self) -> MemoryLock:
        return self._lock

    def _ensure_sheet(self, sheet_name: str) -> None:
        """确认工作表存在；不存在抛 S002。"""
        if sheet_name not in self._workbook.sheetnames:
            raise GXError(
                ERR_STORAGE_SHEET,
                f"工作表不存在: {sheet_name}",
                module="storage",
                context={"sheet_name": sheet_name},
            )
